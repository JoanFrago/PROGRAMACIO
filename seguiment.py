from openpyxl import load_workbook

workbook = load_workbook('SEGUIMENT_SOPAR_ENTRENO.xlsx')

worksheet = workbook.active
num_rows = worksheet.max_row

entrenos = []

for elem in range(1, num_rows + 1):
    cell_object = worksheet.cell(row = elem+1, column = 3).value
    if cell_object == None:
        break
    else:
        entrenos.append(cell_object)
        print(cell_object)
        
print(entrenos)
entrenos.insert(-1, 0)

for elem in entrenos:
    if elem < 7:
        entrenos.remove(elem)
        
print(entrenos)