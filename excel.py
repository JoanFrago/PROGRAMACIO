from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

class Titulo_Primera_Hoja:
    titulo = 'ELEMENTS ESTRUCTURALS DEL COS HUMÀ'   #A1
    
class Column_A:
    info = []
class Column_B:
    info = []
class Column_C:
    info = []
class Column_D:
    info = []
class Column_E:
    info = []
class Column_F:
    info = []
class Column_G:
    info = []
class Column_H:
    info = []
class Column_I:
    info = []
class Column_J:
    info = []

dicc = {1:'A',2:'B',3:'C',4:'D',5:'E',6:'F',7:'G',8:'H',9:'I',10:'J'}

workbook = load_workbook('elements_cos.xlsx')

worksheet = workbook['elements']




for elem in dicc:
    Column_A.info.append(worksheet[str(dicc[elem])+'1'].value)
    print("Column A: ", worksheet[str(dicc[elem])+'1'].value)
for elem in dicc:
    Column_B.info.append(worksheet[str(dicc[elem])+'2'].value)
    print("Column B: ", worksheet[str(dicc[elem])+'2'].value)
for elem in dicc:
    Column_C.info.append(worksheet[str(dicc[elem])+'3'].value)
    print("Column C: ", worksheet[str(dicc[elem])+'3'].value)
for elem in dicc:
    Column_D.info.append(worksheet[str(dicc[elem])+'4'].value)
    print("Column D: ", worksheet[str(dicc[elem])+'4'].value)
for elem in dicc:
    Column_E.info.append(worksheet[str(dicc[elem])+'5'].value)
    print("Column E: ", worksheet[str(dicc[elem])+'5'].value)
for elem in dicc:
    Column_F.info.append(worksheet[str(dicc[elem])+'6'].value)
    print("Column F: ", worksheet[str(dicc[elem])+'6'].value)
for elem in dicc:
    Column_G.info.append(worksheet[str(dicc[elem])+'7'].value)
    print("Column G: ", worksheet[str(dicc[elem])+'7'].value)
for elem in dicc:
    Column_H.info.append(worksheet[str(dicc[elem])+'8'].value)
    print("Column H: ", worksheet[str(dicc[elem])+'8'].value)
for elem in dicc:
    Column_I.info.append(worksheet[str(dicc[elem])+'9'].value)
    print("Column I: ", worksheet[str(dicc[elem])+'9'].value)
for elem in dicc:
    Column_J.info.append(worksheet[str(dicc[elem])+'10'].value)
    print("Column J: ", worksheet[str(dicc[elem])+'10'].value)
    
